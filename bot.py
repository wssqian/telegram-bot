import asyncio
import base64
import ipaddress
import json
import logging
import mimetypes
import os
import re
import sqlite3
import subprocess
import tempfile
import time
from collections import deque
from dataclasses import dataclass
from html import unescape
from pathlib import Path
from typing import Any, Optional
from urllib.parse import quote_plus, urlparse

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from telegram import BotCommand, PhotoSize, Update
from telegram.constants import ChatAction
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)
from telegram.request import HTTPXRequest


load_dotenv()

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO
)
logger = logging.getLogger(__name__)


TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
BOT_PASSWORD = os.getenv("BOT_PASSWORD", "")
AI_API_KEY = os.getenv("AI_API_KEY", "")
AI_API_BASE = os.getenv("AI_API_BASE", "https://api.openai.com/v1").rstrip("/")
AI_MODEL = os.getenv("AI_MODEL", "gpt-4o-mini")
AI_TIMEOUT = int(os.getenv("AI_TIMEOUT", "120"))
AI_TEMPERATURE = float(os.getenv("AI_TEMPERATURE", "0.3"))
AI_MAX_CONTEXT_CHARS = int(os.getenv("AI_MAX_CONTEXT_CHARS", "500"))
AI_MAX_IMAGE_BYTES = int(os.getenv("AI_MAX_IMAGE_BYTES", "350000"))
AI_RETRY_502 = int(os.getenv("AI_RETRY_502", "2"))
AI_BYPASS_PROXY_FOR_LOCAL = os.getenv("AI_BYPASS_PROXY_FOR_LOCAL", "true").lower() in {
    "1",
    "true",
    "yes",
    "on",
}
AI_STREAM = os.getenv("AI_STREAM", "true").lower() in {"1", "true", "yes", "on"}
AI_QUEUE_WORKERS = int(os.getenv("AI_QUEUE_WORKERS", "2"))
AI_MAX_HISTORY_TURNS = int(os.getenv("AI_MAX_HISTORY_TURNS", "8"))
AI_TOOL_MAX_STEPS = int(os.getenv("AI_TOOL_MAX_STEPS", "6"))
AI_ENABLE_TOOLS = os.getenv("AI_ENABLE_TOOLS", "true").lower() in {"1", "true", "yes", "on"}

TELEGRAM_PROXY_URL = os.getenv("TELEGRAM_PROXY_URL", "http://127.0.0.1:7890")
TELEGRAM_POOL_SIZE = int(os.getenv("TELEGRAM_POOL_SIZE", "32"))
TELEGRAM_CONNECT_TIMEOUT = float(os.getenv("TELEGRAM_CONNECT_TIMEOUT", "10"))
TELEGRAM_READ_TIMEOUT = float(os.getenv("TELEGRAM_READ_TIMEOUT", "20"))
TELEGRAM_WRITE_TIMEOUT = float(os.getenv("TELEGRAM_WRITE_TIMEOUT", "20"))

RATE_LIMIT_WINDOW_SECONDS = int(os.getenv("RATE_LIMIT_WINDOW_SECONDS", "30"))
RATE_LIMIT_MAX_REQUESTS = int(os.getenv("RATE_LIMIT_MAX_REQUESTS", "8"))
DB_PATH = os.getenv("DB_PATH", "users.db")
SKILLS_DIR = Path(os.getenv("SKILLS_DIR", "skills"))
FILES_DIR = Path(os.getenv("FILES_DIR", "files"))


BUILTIN_SKILLS: dict[str, str] = {
    "coder": (
        "你是一个编程助理。优先输出可运行代码。"
        "当需要外部信息时可调用 web_search；当需要计算/脚本时可调用 run_python。"
    ),
    "research": (
        "你是一个研究助理。先检索资料，再给结构化结论。"
        "尽量给出来源链接并说明不确定性。"
    ),
    "xlsx-analyst": (
        "你是表格分析助理。优先调用 read_xlsx 查看内容，再做分析。"
        "输出要包含关键数据摘要与结论。"
    ),
}


class AuthStore:
    def __init__(self, db_path: str) -> None:
        self.db_path = db_path
        self._init_db()

    def _connect(self) -> sqlite3.Connection:
        return sqlite3.connect(self.db_path)

    def _init_db(self) -> None:
        with self._connect() as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS users (
                    user_id INTEGER PRIMARY KEY,
                    authorized INTEGER NOT NULL DEFAULT 0
                )
                """
            )
            conn.commit()

    def is_authorized(self, user_id: int) -> bool:
        with self._connect() as conn:
            row = conn.execute(
                "SELECT authorized FROM users WHERE user_id = ?", (user_id,)
            ).fetchone()
        return bool(row and row[0] == 1)

    def ensure_user(self, user_id: int) -> None:
        with self._connect() as conn:
            conn.execute(
                "INSERT OR IGNORE INTO users (user_id, authorized) VALUES (?, 0)",
                (user_id,),
            )
            conn.commit()

    def authorize(self, user_id: int) -> None:
        with self._connect() as conn:
            conn.execute("UPDATE users SET authorized = 1 WHERE user_id = ?", (user_id,))
            conn.commit()


@dataclass
class AIJob:
    mode: str
    payload: dict[str, Any]
    future: asyncio.Future[str]


store = AuthStore(DB_PATH)
ai_queue: asyncio.Queue[AIJob] = asyncio.Queue()
rate_limit_records: dict[int, deque[float]] = {}


def _trim_text(text: str, limit: int = AI_MAX_CONTEXT_CHARS) -> str:
    text = (text or "").strip()
    if len(text) <= limit:
        return text
    return text[: limit - 1] + "…"


def _extract_content(data: dict[str, Any]) -> str:
    choices = data.get("choices", [])
    if not choices:
        return "⚠️ AI 返回为空。"

    msg = choices[0].get("message", {})
    content = msg.get("content", "")
    if isinstance(content, str):
        return content.strip()

    if isinstance(content, list):
        chunks = []
        for part in content:
            if isinstance(part, dict) and part.get("type") == "text":
                text = part.get("text", "")
                if text:
                    chunks.append(text)
        return "\n".join(chunks).strip() or "⚠️ AI 未返回文本内容。"

    return str(content).strip() or "⚠️ AI 未返回有效内容。"


def _is_local_host(hostname: str) -> bool:
    if not hostname:
        return False
    if hostname == "localhost":
        return True
    try:
        ip = ipaddress.ip_address(hostname)
        return ip.is_private or ip.is_loopback
    except ValueError:
        return False


def _get_request_proxies(url: str) -> Optional[dict[str, Optional[str]]]:
    if not AI_BYPASS_PROXY_FOR_LOCAL:
        return None
    parsed = urlparse(url)
    if _is_local_host(parsed.hostname or ""):
        return {"http": None, "https": None}
    return None


def _skill_system_prompt(skill_name: str) -> str:
    if not skill_name:
        return ""
    skill_file = SKILLS_DIR / f"{skill_name}.txt"
    if not skill_file.exists():
        return ""
    try:
        return skill_file.read_text(encoding="utf-8").strip()
    except Exception:
        return ""


def _build_text_messages(
    user_text: str, history: list[dict[str, str]], skill_name: str = ""
) -> list[dict[str, Any]]:
    base_prompt = "你是一个有帮助的 Telegram 助手，请使用简体中文回答。"
    skill_prompt = _skill_system_prompt(skill_name)
    system_content = base_prompt if not skill_prompt else f"{base_prompt}\n\n当前技能({skill_name})：{skill_prompt}"

    messages: list[dict[str, Any]] = [{"role": "system", "content": system_content}]
    for item in history[-AI_MAX_HISTORY_TURNS * 2 :]:
        role = item.get("role")
        content = item.get("content", "")
        if role in {"user", "assistant"} and content:
            messages.append({"role": role, "content": _trim_text(content, 1600)})
    messages.append({"role": "user", "content": _trim_text(user_text, 1600)})
    return messages


def _build_vision_messages(
    image_bytes: bytes,
    mime_type: str,
    prev_text: str,
    next_text: str,
    caption: str,
) -> list[dict[str, Any]]:
    image_b64 = base64.b64encode(image_bytes).decode("utf-8")
    image_data_url = f"data:{mime_type};base64,{image_b64}"
    instruction = (
        "请做图片视觉理解，回答简洁。\n"
        f"上文：{_trim_text(prev_text)}\n"
        f"下文：{_trim_text(next_text)}\n"
        f"备注：{_trim_text(caption)}\n"
        "输出格式：\n1) 画面内容\n2) 结合上下文的回复"
    )
    return [
        {"role": "system", "content": "你是支持视觉理解的中文助手。"},
        {
            "role": "user",
            "content": [
                {"type": "text", "text": instruction},
                {"type": "image_url", "image_url": {"url": image_data_url}},
            ],
        },
    ]


def _chat_request(payload: dict[str, Any], *, stream: bool = False):
    url = f"{AI_API_BASE}/chat/completions"
    headers = {
        "Authorization": f"Bearer {AI_API_KEY}",
        "Content-Type": "application/json",
    }
    return requests.post(
        url,
        headers=headers,
        json=payload,
        timeout=AI_TIMEOUT,
        proxies=_get_request_proxies(url),
        stream=stream,
    )


def _stream_chat_completions(messages: list[dict[str, Any]]) -> str:
    payload: dict[str, Any] = {
        "model": AI_MODEL,
        "messages": messages,
        "temperature": AI_TEMPERATURE,
        "stream": True,
    }
    attempts = max(1, AI_RETRY_502 + 1)
    for idx in range(attempts):
        try:
            chunks: list[str] = []
            with _chat_request(payload, stream=True) as resp:
                resp.raise_for_status()
                for raw_line in resp.iter_lines(decode_unicode=False):
                    if not raw_line:
                        continue
                    line = raw_line.decode("utf-8", errors="ignore").strip()
                    if not line.startswith("data:"):
                        continue
                    data = line[5:].strip()
                    if data == "[DONE]":
                        break
                    try:
                        obj = json.loads(data)
                    except ValueError:
                        continue
                    delta = obj.get("choices", [{}])[0].get("delta", {})
                    token = delta.get("content")
                    if token:
                        chunks.append(token)
                final_text = "".join(chunks).strip()
                if final_text:
                    return final_text
                return "⚠️ AI 流式返回为空。"
        except requests.HTTPError as exc:
            status = exc.response.status_code if exc.response is not None else None
            if status in {502, 503, 504} and idx < attempts - 1:
                time.sleep(min(2.5, 0.8 * (idx + 1)))
                continue
            detail = exc.response.text[:240] if exc.response is not None else ""
            logger.exception("AI stream HTTP error: %s", exc)
            return f"⚠️ 调用 AI 失败：HTTP {status or 'unknown'} {detail}"
        except requests.RequestException as exc:
            logger.exception("AI stream request failed: %s", exc)
            return f"⚠️ 调用 AI 失败：{exc}"
    return "⚠️ 调用 AI 失败：网关繁忙，请稍后重试。"


def _call_chat_completions(messages: list[dict[str, Any]]) -> str:
    if not AI_API_KEY:
        return "⚠️ 未配置 AI_API_KEY，暂时无法调用 AI。"

    if AI_STREAM:
        streamed = _stream_chat_completions(messages)
        if not streamed.startswith("⚠️ 调用 AI 失败"):
            return streamed

    payload: dict[str, Any] = {
        "model": AI_MODEL,
        "messages": messages,
        "temperature": AI_TEMPERATURE,
    }

    attempts = max(1, AI_RETRY_502 + 1)
    for idx in range(attempts):
        try:
            resp = _chat_request(payload)
            resp.raise_for_status()
            return _extract_content(resp.json())
        except requests.HTTPError as exc:
            status = exc.response.status_code if exc.response is not None else None
            if status in {502, 503, 504} and idx < attempts - 1:
                time.sleep(min(2.5, 0.8 * (idx + 1)))
                continue
            detail = exc.response.text[:240] if exc.response is not None else ""
            logger.exception("AI API HTTP error: %s", exc)
            return f"⚠️ 调用 AI 失败：HTTP {status or 'unknown'} {detail}"
        except requests.RequestException as exc:
            logger.exception("AI API request failed: %s", exc)
            return f"⚠️ 调用 AI 失败：{exc}"
    return "⚠️ 调用 AI 失败：网关繁忙，请稍后重试。"


def _tool_web_search(query: str, max_results: int = 5) -> str:
    max_results = max(1, min(max_results, 10))
    url = f"https://duckduckgo.com/html/?q={quote_plus(query)}"
    try:
        resp = requests.get(url, timeout=20)
        resp.raise_for_status()
        html = resp.text
        matches = re.findall(r'<a[^>]+class="result__a"[^>]+href="([^"]+)"[^>]*>(.*?)</a>', html)
        out = []
        for idx, (link, title_html) in enumerate(matches[:max_results], 1):
            title = re.sub(r"<.*?>", "", unescape(title_html)).strip()
            out.append(f"{idx}. {title}\n{unescape(link)}")
        return "\n\n".join(out) if out else "未检索到结果。"
    except Exception as exc:  # noqa: BLE001
        return f"搜索失败：{exc}"


def _tool_read_xlsx(path: str, sheet_name: str = "", max_rows: int = 30) -> str:
    target = (FILES_DIR / path).resolve() if not Path(path).is_absolute() else Path(path).resolve()
    if not str(target).startswith(str(FILES_DIR.resolve())) and not Path(path).is_absolute():
        return "路径非法，仅允许读取 FILES_DIR 下文件。"
    if not target.exists():
        return f"文件不存在：{target}"
    if target.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return "文件不是 xlsx 格式。"

    wb = load_workbook(target, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    lines = [f"sheet={ws.title}"]
    for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if idx > max_rows:
            lines.append("...已截断...")
            break
        vals = ["" if v is None else str(v) for v in row]
        lines.append("\t".join(vals))
    return "\n".join(lines)


def _tool_list_files(subdir: str = ".", max_items: int = 100) -> str:
    base = (FILES_DIR / subdir).resolve()
    if not str(base).startswith(str(FILES_DIR.resolve())):
        return "路径非法。"
    if not base.exists():
        return "目录不存在。"
    items = []
    for p in sorted(base.iterdir())[:max_items]:
        items.append(("[D]" if p.is_dir() else "[F]") + f" {p.name}")
    return "\n".join(items) if items else "目录为空。"


def _tool_read_text(path: str, max_chars: int = 4000) -> str:
    target = (FILES_DIR / path).resolve() if not Path(path).is_absolute() else Path(path).resolve()
    if not str(target).startswith(str(FILES_DIR.resolve())) and not Path(path).is_absolute():
        return "路径非法，仅允许读取 FILES_DIR 下文件。"
    if not target.exists():
        return f"文件不存在：{target}"
    try:
        return target.read_text(encoding="utf-8", errors="ignore")[:max_chars]
    except Exception as exc:  # noqa: BLE001
        return f"读取失败：{exc}"


def _tool_run_python(code: str, timeout_s: int = 8) -> str:
    timeout_s = max(1, min(timeout_s, 20))
    try:
        proc = subprocess.run(
            ["python", "-c", code],
            capture_output=True,
            text=True,
            timeout=timeout_s,
        )
        out = (proc.stdout or "")[:6000]
        err = (proc.stderr or "")[:3000]
        return f"exit={proc.returncode}\nstdout:\n{out}\nstderr:\n{err}"
    except Exception as exc:  # noqa: BLE001
        return f"执行失败：{exc}"


def _tool_specs() -> list[dict[str, Any]]:
    return [
        {
            "type": "function",
            "function": {
                "name": "web_search",
                "description": "联网搜索网页结果（DuckDuckGo）",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "query": {"type": "string"},
                        "max_results": {"type": "integer", "default": 5},
                    },
                    "required": ["query"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "list_files",
                "description": "列出 FILES_DIR 下目录内容",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "subdir": {"type": "string", "default": "."},
                        "max_items": {"type": "integer", "default": 100},
                    },
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "read_text",
                "description": "读取文本文件",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "max_chars": {"type": "integer", "default": 4000},
                    },
                    "required": ["path"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "read_xlsx",
                "description": "读取 xlsx 文件并输出前若干行",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string"},
                        "sheet_name": {"type": "string", "default": ""},
                        "max_rows": {"type": "integer", "default": 30},
                    },
                    "required": ["path"],
                },
            },
        },
        {
            "type": "function",
            "function": {
                "name": "run_python",
                "description": "执行短 Python 代码片段",
                "parameters": {
                    "type": "object",
                    "properties": {
                        "code": {"type": "string"},
                        "timeout_s": {"type": "integer", "default": 8},
                    },
                    "required": ["code"],
                },
            },
        },
    ]


def _exec_tool(name: str, args: dict[str, Any]) -> str:
    if name == "web_search":
        return _tool_web_search(args.get("query", ""), int(args.get("max_results", 5)))
    if name == "list_files":
        return _tool_list_files(args.get("subdir", "."), int(args.get("max_items", 100)))
    if name == "read_text":
        return _tool_read_text(args.get("path", ""), int(args.get("max_chars", 4000)))
    if name == "read_xlsx":
        return _tool_read_xlsx(
            args.get("path", ""),
            args.get("sheet_name", ""),
            int(args.get("max_rows", 30)),
        )
    if name == "run_python":
        return _tool_run_python(args.get("code", ""), int(args.get("timeout_s", 8)))
    return f"未知工具：{name}"


def _call_chat_with_tools(messages: list[dict[str, Any]]) -> str:
    if not AI_API_KEY:
        return "⚠️ 未配置 AI_API_KEY，暂时无法调用 AI。"

    tool_messages = list(messages)
    for _ in range(max(1, AI_TOOL_MAX_STEPS)):
        payload = {
            "model": AI_MODEL,
            "messages": tool_messages,
            "temperature": AI_TEMPERATURE,
            "tools": _tool_specs(),
            "tool_choice": "auto",
        }
        try:
            resp = _chat_request(payload)
            resp.raise_for_status()
            data = resp.json()
        except requests.RequestException as exc:
            return f"⚠️ 调用 AI 失败：{exc}"

        choices = data.get("choices", [])
        if not choices:
            return "⚠️ AI 返回为空。"

        msg = choices[0].get("message", {})
        tool_calls = msg.get("tool_calls") or []
        if not tool_calls:
            return _extract_content(data)

        assistant_msg = {
            "role": "assistant",
            "content": msg.get("content", ""),
            "tool_calls": tool_calls,
        }
        tool_messages.append(assistant_msg)

        for tc in tool_calls:
            fn = tc.get("function", {})
            name = fn.get("name", "")
            raw_args = fn.get("arguments", "{}")
            try:
                args = json.loads(raw_args) if raw_args else {}
            except json.JSONDecodeError:
                args = {}
            result = _exec_tool(name, args)
            tool_messages.append(
                {
                    "role": "tool",
                    "tool_call_id": tc.get("id", ""),
                    "name": name,
                    "content": _trim_text(result, 12000),
                }
            )

    return "⚠️ 工具调用步数超限，请缩小问题范围后重试。"


async def enqueue_ai_job(mode: str, payload: dict[str, Any]) -> str:
    loop = asyncio.get_running_loop()
    future: asyncio.Future[str] = loop.create_future()
    await ai_queue.put(AIJob(mode=mode, payload=payload, future=future))
    return await future


async def ai_worker() -> None:
    while True:
        job = await ai_queue.get()
        try:
            if job.mode == "text":
                messages = _build_text_messages(
                    job.payload.get("text", ""),
                    job.payload.get("history", []),
                    job.payload.get("skill_name", ""),
                )
                if AI_ENABLE_TOOLS:
                    result = await asyncio.to_thread(_call_chat_with_tools, messages)
                else:
                    result = await asyncio.to_thread(_call_chat_completions, messages)
            else:
                messages = _build_vision_messages(
                    job.payload["image_bytes"],
                    job.payload["mime_type"],
                    job.payload.get("prev_text", ""),
                    job.payload.get("next_text", ""),
                    job.payload.get("caption", ""),
                )
                result = await asyncio.to_thread(_call_chat_completions, messages)
            if not job.future.done():
                job.future.set_result(result)
        except Exception as exc:  # noqa: BLE001
            logger.exception("AI worker error: %s", exc)
            if not job.future.done():
                job.future.set_result(f"⚠️ AI 处理失败：{exc}")
        finally:
            ai_queue.task_done()


def is_authorized(user_id: Optional[int]) -> bool:
    if user_id is None:
        return False
    store.ensure_user(user_id)
    return store.is_authorized(user_id)


def allow_user_request(user_id: int) -> bool:
    now = time.time()
    q = rate_limit_records.setdefault(user_id, deque())
    while q and now - q[0] > RATE_LIMIT_WINDOW_SECONDS:
        q.popleft()
    if len(q) >= RATE_LIMIT_MAX_REQUESTS:
        return False
    q.append(now)
    return True


def _pick_reasonable_photo(photos: list[PhotoSize]) -> PhotoSize:
    sorted_items = sorted(photos, key=lambda p: p.file_size or 0)
    for item in sorted_items:
        if (item.file_size or 0) <= AI_MAX_IMAGE_BYTES:
            return item
    return sorted_items[0]


def _reset_chat_context(context: ContextTypes.DEFAULT_TYPE) -> None:
    context.user_data.pop("pending_photo", None)
    context.user_data["history"] = []


def _ensure_dirs() -> None:
    SKILLS_DIR.mkdir(parents=True, exist_ok=True)
    FILES_DIR.mkdir(parents=True, exist_ok=True)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    user_id = update.effective_user.id if update.effective_user else None
    if user_id is None or update.message is None:
        return

    store.ensure_user(user_id)
    context.user_data.setdefault("history", [])
    if store.is_authorized(user_id):
        await update.message.reply_text("欢迎回来！你已经通过验证。\n发送文字、图片或文件，我都会处理。")
    else:
        await update.message.reply_text("欢迎！首次使用请先输入访问密码。")


async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None:
        return
    await update.message.reply_text(
        "使用说明：\n"
        "1) 首次聊天需要输入密码认证。\n"
        "2) 支持多轮上下文对话。\n"
        "3) /new 与 /refresh 可清空上下文。\n"
        "4) 支持技能：/skill_list /skill_install <name> /skill_use <name> /skill_off。\n"
        "5) 工具能力：联网搜索、读取 xlsx、读取文件、执行 Python。"
    )


async def new_chat_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None:
        return
    _reset_chat_context(context)
    await update.message.reply_text("✅ 已开启新对话，上下文已清空。")


async def refresh_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None:
        return
    _reset_chat_context(context)
    await update.message.reply_text("🔄 聊天已刷新，上下文已清空。")


async def skill_list_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None:
        return
    _ensure_dirs()
    names = sorted(p.stem for p in SKILLS_DIR.glob("*.txt"))
    builtins = ", ".join(sorted(BUILTIN_SKILLS.keys()))
    current = context.user_data.get("skill_name", "无")
    text = "\n".join(names) if names else "(无已安装技能)"
    await update.message.reply_text(
        f"当前技能：{current}\n\n已安装技能：\n{text}\n\n可安装内置技能：{builtins}"
    )


async def skill_install_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None:
        return
    _ensure_dirs()
    if not context.args:
        await update.message.reply_text("用法：/skill_install coder|research|xlsx-analyst")
        return
    name = context.args[0].strip().lower()
    prompt = BUILTIN_SKILLS.get(name)
    if not prompt:
        await update.message.reply_text("内置技能不存在，可用：coder, research, xlsx-analyst")
        return
    (SKILLS_DIR / f"{name}.txt").write_text(prompt, encoding="utf-8")
    await update.message.reply_text(f"✅ 已安装技能：{name}")


async def skill_use_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None:
        return
    if not context.args:
        await update.message.reply_text("用法：/skill_use <skill_name>")
        return
    name = context.args[0].strip()
    if not (SKILLS_DIR / f"{name}.txt").exists():
        await update.message.reply_text(f"技能不存在：{name}，先 /skill_install 或手动放入 {SKILLS_DIR}")
        return
    context.user_data["skill_name"] = name
    await update.message.reply_text(f"✅ 已启用技能：{name}")


async def skill_off_command(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None:
        return
    context.user_data.pop("skill_name", None)
    await update.message.reply_text("已关闭技能，恢复默认助手模式。")


async def handle_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None:
        return

    user_id = update.effective_user.id if update.effective_user else None
    text = (update.message.text or "").strip()
    if user_id is None:
        return

    store.ensure_user(user_id)
    context.user_data.setdefault("history", [])

    if not store.is_authorized(user_id):
        if not BOT_PASSWORD:
            await update.message.reply_text("⚠️ 服务端未设置 BOT_PASSWORD，无法完成认证。")
            return
        if text == BOT_PASSWORD:
            store.authorize(user_id)
            await update.message.reply_text("✅ 密码正确，认证成功！现在可以开始聊天了。")
        else:
            await update.message.reply_text("❌ 密码错误，请重新输入。")
        return

    if not allow_user_request(user_id):
        await update.message.reply_text("⚠️ 你发送太快了，请稍后再试。")
        return

    pending_photo = context.user_data.get("pending_photo")
    await context.bot.send_chat_action(chat_id=update.effective_chat.id, action=ChatAction.TYPING)

    if pending_photo:
        context.user_data["pending_photo"] = None
        ai_reply = await enqueue_ai_job(
            "vision",
            {
                "image_bytes": pending_photo["image_bytes"],
                "mime_type": pending_photo["mime_type"],
                "prev_text": pending_photo.get("prev_text", ""),
                "next_text": text,
                "caption": pending_photo.get("caption", ""),
            },
        )
        await update.message.reply_text(f"已收到你的图片。\nAI：{ai_reply}")
        return

    placeholder = await update.message.reply_text("⏳ AI 正在回复中...")
    history = context.user_data.get("history", [])
    skill_name = context.user_data.get("skill_name", "")
    ai_reply = await enqueue_ai_job(
        "text", {"text": text, "history": history, "skill_name": skill_name}
    )

    if ai_reply and not ai_reply.startswith("⚠️"):
        history.append({"role": "user", "content": text})
        history.append({"role": "assistant", "content": ai_reply})
        context.user_data["history"] = history[-AI_MAX_HISTORY_TURNS * 2 :]

    chunks = [ai_reply[i : i + 700] for i in range(0, len(ai_reply), 700)] or [ai_reply]
    first = True
    for idx, chunk in enumerate(chunks):
        if first:
            await placeholder.edit_text(chunk)
            first = False
        else:
            await update.message.reply_text(chunk)
        if idx < len(chunks) - 1:
            await asyncio.sleep(0.15)


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None:
        return

    user_id = update.effective_user.id if update.effective_user else None
    if user_id is None or not is_authorized(user_id):
        await update.message.reply_text("请先输入密码完成认证，再发送图片。")
        return

    if not allow_user_request(user_id):
        await update.message.reply_text("⚠️ 你发送太快了，请稍后再试。")
        return

    selected = _pick_reasonable_photo(update.message.photo)
    tg_file = await selected.get_file()
    image_bytes = bytes(await tg_file.download_as_bytearray())

    if len(image_bytes) > AI_MAX_IMAGE_BYTES:
        await update.message.reply_text("图片较大，已超出当前网关限制，请发送更低分辨率图片后重试。")
        return

    caption = _trim_text(update.message.caption or "", 300)
    history = context.user_data.get("history", [])
    prev_text = history[-1]["content"] if history and history[-1].get("role") == "assistant" else ""

    mime_type = "image/jpeg"
    if tg_file.file_path:
        guessed_mime, _ = mimetypes.guess_type(tg_file.file_path)
        if guessed_mime:
            mime_type = guessed_mime

    if caption:
        await context.bot.send_chat_action(chat_id=update.effective_chat.id, action=ChatAction.TYPING)
        ai_reply = await enqueue_ai_job(
            "vision",
            {
                "image_bytes": image_bytes,
                "mime_type": mime_type,
                "prev_text": prev_text,
                "next_text": caption,
                "caption": caption,
            },
        )
        await update.message.reply_text(f"已收到你的图片。\nAI：{ai_reply}")
    else:
        context.user_data["pending_photo"] = {
            "image_bytes": image_bytes,
            "mime_type": mime_type,
            "prev_text": prev_text,
            "caption": caption,
        }
        await update.message.reply_text("已收到你的图片。请再发送一句补充描述，我会做视觉联合分析。")


async def handle_document(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if update.message is None:
        return

    user_id = update.effective_user.id if update.effective_user else None
    if user_id is None or not is_authorized(user_id):
        await update.message.reply_text("请先输入密码完成认证，再发送文件。")
        return

    if not allow_user_request(user_id):
        await update.message.reply_text("⚠️ 你发送太快了，请稍后再试。")
        return

    document = update.message.document
    if document is None:
        return

    tg_file = await document.get_file()
    with tempfile.TemporaryDirectory() as tmp_dir:
        filename = document.file_name or f"doc_{document.file_unique_id}"
        local_path = Path(tmp_dir) / filename
        await tg_file.download_to_drive(custom_path=str(local_path))

        prompt = (
            f"用户上传了文件：{filename}，大小约 {document.file_size or 0} 字节。"
            "请提示用户把文件放到 FILES_DIR 后可用 read_text/read_xlsx 工具读取分析。"
        )
        ai_reply = await enqueue_ai_job("text", {"text": prompt, "history": []})

        with local_path.open("rb") as f:
            await update.message.reply_document(
                document=f,
                caption=f"已收到并回传文件：{filename}\nAI：{ai_reply}",
            )


async def post_init(app: Application) -> None:
    _ensure_dirs()

    workers: list[asyncio.Task[Any]] = []
    for _ in range(max(1, AI_QUEUE_WORKERS)):
        workers.append(asyncio.create_task(ai_worker()))
    app.bot_data["ai_workers"] = workers

    await app.bot.set_my_commands(
        [
            BotCommand("start", "开始使用"),
            BotCommand("help", "查看帮助"),
            BotCommand("new", "开启新对话"),
            BotCommand("refresh", "刷新并清空上下文"),
            BotCommand("skill_list", "查看技能"),
            BotCommand("skill_install", "安装内置技能"),
            BotCommand("skill_use", "启用技能"),
            BotCommand("skill_off", "关闭技能"),
        ]
    )


async def post_shutdown(app: Application) -> None:
    workers = app.bot_data.get("ai_workers", [])
    for task in workers:
        task.cancel()
    if workers:
        await asyncio.gather(*workers, return_exceptions=True)


def validate_env() -> None:
    if not TELEGRAM_BOT_TOKEN:
        raise RuntimeError("缺少 TELEGRAM_BOT_TOKEN 环境变量")


def main() -> None:
    validate_env()

    request = HTTPXRequest(
        connection_pool_size=TELEGRAM_POOL_SIZE,
        proxy_url=TELEGRAM_PROXY_URL or None,
        connect_timeout=TELEGRAM_CONNECT_TIMEOUT,
        read_timeout=TELEGRAM_READ_TIMEOUT,
        write_timeout=TELEGRAM_WRITE_TIMEOUT,
    )

    app = (
        Application.builder()
        .token(TELEGRAM_BOT_TOKEN)
        .request(request)
        .get_updates_request(request)
        .post_init(post_init)
        .post_shutdown(post_shutdown)
        .build()
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("help", help_command))
    app.add_handler(CommandHandler("new", new_chat_command))
    app.add_handler(CommandHandler("refresh", refresh_command))
    app.add_handler(CommandHandler("Refresh", refresh_command))
    app.add_handler(CommandHandler("skill_list", skill_list_command))
    app.add_handler(CommandHandler("skill_install", skill_install_command))
    app.add_handler(CommandHandler("skill_use", skill_use_command))
    app.add_handler(CommandHandler("skill_off", skill_off_command))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_text))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    app.add_handler(MessageHandler(filters.Document.ALL, handle_document))

    logger.info("Bot started.")
    app.run_polling(close_loop=False)


if __name__ == "__main__":
    main()
